# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 09:44:10 2017

@author: astachn1

Unfortunately, DePaul's security policies prevent this from being as good
a script as it could be. Specifically, DePaul sets the McAfee Virus Scan to
default block all attempts to access port 587 for smtp email, meaning it's
impossible to even connect to the DPU server (or any server). There's no way
to change the McAfee settings either, as they're locked down. Having Python
directly control Outlook through the win32 client used to work very well,
but they've recently added protection which now requires you to press "Allow"
every time the program attempts to send an email. That security setting is
also locked. So, this script suffers.
"""

import os
from datetime import datetime
import pandas as pd
import win32com.client as win32
import re
from time import sleep
import errno
import pyodbc
import click
import openpyxl
import dpu.scripts as dpu
from dpu.file_locator import FileLocator

#############################################################
# Functions
#############################################################
def send_email (recipient, subject, body, attachments):
    '''Sends email using the Outlook client'''
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.Body = body
    if attachments:
        for attachment in attachments:
            mail.Attachments.Add(Source = attachment)
    mail.Send()

def email_preceptor (row):
    '''Emails each preceptor in the file.'''
    # Set save path
    FL = FileLocator()
    ending_path = FL.contracts
    #Read in term descriptions
    TermDescriptions = dpu.get_term_descriptions()
    # Set path for attachment
    FL.preceptor
    duties = os.path.abspath(os.path.join(os.sep, FL.preceptor, 'MENP', 'Preceptor, Student, and Faculty Responsibilities for MENP Program.pdf'))
    
    # Get coordinator's email
    recipient = row['Preceptor Email']
    # Write subject
    subject = 'DePaul University Preceptorship'
    # Write body   
    body = '''To: {0} {1}
    Re:	DePaul University Preceptorship

    Dear {2}:
    Thank you for agreeing to serve as a Preceptor for DePaul University student {3} {4} (the "Student"), for the time period {5}. 
    
    As a Preceptor you will provide a supervised learning experience for the Student, who will be simultaneously enrolled in the course NSG {6}-{7}: {8} (the "Course"). The Student's specific objectives for the learning experience and the Course will be jointly planned and approved by the University, the Preceptor, and the Student at the initiation of the learning experience. The learning experience must provide the Student with a minimum of {9} hours of direct, supervised clinical practice consistent with the State of Illinois Nurse Practice Act and in keeping with your typical responsibilities in your work setting. Evaluating the Student's achievement of the approved objectives will be a joint responsibility of the University, the Preceptor, and the Student at the conclusion of the learning experience.
    
    Your specific responsibilities as Preceptor are outlined in the attached document, entitled "Preceptor, Student, and Faculty Responsibilities."  We would also encourage you to review the Affiliation Agreement between DePaul University and your site for more information about the rights and responsibilities of various parties with respect to the placement. Please be sure to share this letter with your immediate supervisor. In the event that you cannot fulfill these responsibilities and your immediate supervisor assigns another preceptor for the Student, this letter applies to anyone who may be assigned as preceptor for the Student.
    
    I again thank you for agreeing to participate in this program and provide an invaluable learning experience for {10} {11}.  If you have any questions or concerns, please do not hesitate to contact me.
    
    Sincerely,
    
    DePaul University
    School of Nursing'''.format(row['Preceptor First Name'], row['Preceptor Last Name'], row['Preceptor First Name'], row['Student First Name'], row['Student Last Name'], row['Date'], row['Cr'], row['Sec'], row['Title'], row['Hours'], row['Student First Name'], row['Student Last Name'])
    
    # Get attachments    
    attachments = [duties]
    # Send email
    try:
        send_email(recipient, subject, body, attachments)
        # If successful, add email to list of jobs completed
        jobs_completed.append(recipient)
    except Exception as e:
        print('Exception occured for {0}: {1}'.format(recipient, e))
    
    # Map clinical site to official organization name
    # Gather sites and abbreviations
    site_dict = get_sites_from_ARC()
    # Convert to a list of all sites
    sites = list(site_dict.keys())
    # Gather dictionary of site abbreviations for storage/lookup
    abbr_dict = {}
    for site in sites:
        # Convert site name to an abbreviation
        abbr = re.sub(r'[^A-Z]', '', site)
        # If duplicate abbreviations exist, store as a list
        if abbr in abbr_dict:
            if type(abbr_dict[abbr]) == str:
                abbr_dict[abbr] = [abbr_dict[abbr]]
            abbr_dict[abbr].append(site)
        # Otherwise, store as a single string
        else:
            abbr_dict[abbr] = site
    abbr = list(abbr_dict.keys())
    # Test if org is in the site_dict
    if row['Clinical Site'] in site_dict.keys():
        org = site_dict[row['Clinical Site']]
    else:
        # Test if org is an abbreviation
        if re.match(r'^[A-Z]+$', row['Clinical Site']):
            org = dpu.find_best_string_match(row['Clinical Site'], abbr)
            # It is possible for multiple sites to share an abbreviation
            # If this happens, we can attempt to use the recipient email
            email_org = re.findall(r'@(\w+)\.', recipient)[0]
            org = dpu.find_best_string_match(email_org, org)
        else:
            org = dpu.find_best_string_match(row['Clinical Site'], sites)

    # Save a copy of the email
    save_ay = TermDescriptions.loc[TermDescriptions['Term'] == row['Term'], 'Academic Year'].item()
    save_q = TermDescriptions.loc[TermDescriptions['Term'] == row['Term'], 'Quarter'].item()
    save_path = os.path.abspath(os.path.join(os.sep, ending_path, org, 'Preceptor Letters of Agreement', save_ay, save_q))
    
    title = row['Preceptor Last Name'] + ', ' + row['Preceptor First Name']
    # Sleep for 1 second to allow time for Outlook to refresh
    sleep(1)
    save_email('Preceptor', save_path, title)

def test_path (save_path):
    '''Creates path. If exception = Path Exists, ignore. Otherwise raise.'''
    try:
        os.makedirs(save_path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

def save_email (folder, save_path, title):
    '''Saves the last email in a particular folder.
    Additional documentation: https://msdn.microsoft.com/en-us/library/microsoft.office.interop.outlook.mailitem_methods.aspx
    '''
    email_address = os.getlogin() + "@depaul.edu"
    outlook = win32.Dispatch('outlook.application').GetNamespace("MAPI")
    inbox = outlook.Folders(email_address).Folders(folder)
    messages = inbox.Items
    message = messages.GetLast()
    test_path(save_path)
    message.SaveAs(save_path + '\\' + title + '.msg')
    
def get_sites_from_ARC ():
    '''A function that connects to the ARC Access database and grabs all
    site / corporation pairings.'''
    FL = FileLocator()
    # Access Connection Strings
    access_conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ={0};'.format(os.path.abspath(os.path.join(os.sep, FL.arc, 'Backend', 'AffiliationAgreements_Backend.accdb')))
        )
    # Connect and open cursor
    cnxn = pyodbc.connect(access_conn_str)
    crsr = cnxn.cursor()
    # multiple join SQL statement
    sql = "SELECT c.[Site Name], d.Corporation FROM ((([Sites in Agreement] AS a) INNER JOIN Agreements AS b ON b.Agreement_ID = a.Agreement) INNER JOIN Sites AS c ON c.Site_ID = a.Site) INNER JOIN Corporations AS d ON d.Corporation_ID = b.Corporation"
    # Execute and save query results to dictionary
    crsr.execute(sql)
    site_dict = dict(crsr.fetchall())
    # Close connection
    crsr.close()
    cnxn.close()
    return site_dict

def save_excel_changes (file):
    '''Function saves changes made to excel file'''
    # Open excel workbook
    wb = openpyxl.load_workbook(file)
    # Open first sheet
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    # Gather column indices
    col_indices = {cell.value:n for n, cell in enumerate(list(sheet.rows)[0])}
    # Find column for Preceptor Email and Letter Sent
    colNum_email = col_indices['Preceptor Email'] + 1
    colNum_sent = col_indices['Preceptor Letter Sent'] + 1
    # Get today's date
    today = datetime.today().strftime('%m/%d/%Y')
    # Iterate through all rows in sheet, skipping first row (idx starts at 1)
    for rowNum in range(2, sheet.max_row):
        # Test if the email is in the jobs_completed list
        if sheet.cell(row=rowNum, column=colNum_email).value in jobs_completed:
            # Write today's date
            sheet.cell(row=rowNum, column=colNum_sent).value = today
    # Save workbook
    wb.save(file)

#############################################################
# Main
#############################################################
@click.command()
@click.option(
        '--term', '-t', type=str,
        help='Term for which to send out preceptor emails',
)
def main(term):
    '''Main function.'''
    # Make a global list to store jobs completed
    global jobs_completed
    jobs_completed = []
    
    # If user did not pass a term, guess current
    if not term:
        term = dpu.guess_current_term()
    # Get preceptors
    preceptors, file = dpu.get_cln(term, dpu.get_term_descriptions())
    # Drop any rows without a preceptor email
    preceptors.dropna(subset=['Preceptor Email'], inplace=True)
    # Drop any rows with "TBD" as preceptor email
    preceptors = preceptors[~preceptors['Preceptor Email'].isin(['TBD', 'TBA'])]
    # Drop any rows where preceptor email was already sent or will be sent by RFU
    preceptors = preceptors[preceptors['Preceptor Letter Sent'].isnull()]
    # Email the preceptors
    preceptors.apply(email_preceptor, axis=1)
    # Save the current date in Excel file for preceptor's emailed
    save_excel_changes (file)

if __name__ == "__main__":
    main()
