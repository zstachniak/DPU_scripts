# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 14:49:40 2017

@author: astachn1
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 09:44:10 2017

@author: astachn1
"""

import os
from datetime import datetime
import pandas as pd
import win32com.client as win32
import re
from time import sleep
import errno
import pyodbc
from fuzzywuzzy import fuzz, process
import click
import openpyxl
from openpyxl.styles import NamedStyle

#############################################################
# Functions
#############################################################
def find_best_string_match (query, choices, **kwargs):
    '''This function takes a single query and a list of possible
    choices and ranks the choices to find the most likely match.
    Rankings are calculated via fuzzywuzzy ratios, and can be
    passed directly by the user via optional keyword.'''
    # Optional argument to test only certain scorers
    scorers = kwargs.pop('scorers', [fuzz.ratio, fuzz.partial_ratio, fuzz.token_sort_ratio, fuzz.token_set_ratio])
    # Initialize a dictionary to store scoring
    score_mapping = {}
    for key in choices:
        score_mapping[key] = []
    # Test for each scorer
    for scorer in scorers:
        # Store temporary results as list of tuples
        temp_results = process.extract(query, choices, scorer=scorer, limit=None)
        # Add scores to mapping
        for (key, score) in temp_results:
            score_mapping[key].append(score)
    # Sum all results for each key
    for key in score_mapping.keys():
        score_mapping[key] = sum(score_mapping[key])
    # Determine the maximum scored
    result = max(score_mapping, key=lambda key: score_mapping[key])
    return result

def get_dir_paths (pathname, pattern, ignore_list):
    '''Given a starting pathname and a pattern, function will return
    and subdirectories that match the regex pattern.'''
    dirs = []
    # Iterate through subfiles
    for name in os.listdir(pathname):
        subname = os.path.join(pathname, name)
        # Checks if subpath is a directory and that it matches pattern
        if os.path.isdir(subname) and re.fullmatch(pattern, name) and name not in ignore_list:
            dirs.append(subname)
    return dirs

def get_latest (pathname):
    '''Scans the folder for all files that match typical naming conventions.
    For those files, function parses file name to look for date edited,
    then returns the file name with the latest date.'''
    
    #print('Scanning: {}'.format(pathname))
    
    #Set up empty lists
    files = []
    dates = []
    
    for name in os.listdir(pathname):
        subname = os.path.join(pathname, name)
        
        #Checks for standard naming conventions and ignores file fragments
        if os.path.isfile(subname) and 'Clinical Roster' in subname and '~$' not in subname:
            #Ignore files that end in '_(2).xlsx'
            if (subname[(subname.find('.')-3):subname.find('.')]) == '(2)':
                pass
            else:
                files.append(subname)
    
    #If only one file, return that one
    if len(files) == 1:
        return files[0]
    
    #If multiple files, return the one that contains the latest date
    else:
        #Parses file names and converts to datetimes
        for subname in files:
            date = subname[(subname.find('.')-10):subname.find('.')]
            date_time = datetime.strptime(date, '%Y-%m-%d').date()
            dates.append(date_time)
            #print('Adding: {}'.format(subname))
        latest = max(dates)
        #print(latest.strftime('%m-%d-%Y'))
        for file in files:
            if str(latest.strftime('%Y-%m-%d')) in file:
                return file

def remove_files (folder):
    '''Removes all files in a given folder.'''
    for file in os.listdir(folder):
        file_path = os.path.join(folder, file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(e)

def send_email (recipient, subject, body, attachments):
    '''Sends email using the Outlook client'''
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.Body = body
    if attachments:
        for attachment in attachments:
            mail.Attachments.Add(Source = attachment)
    mail.Send()

def email_preceptor (row):
    '''Emails each preceptor in the file.'''
    # Set save path
    ending_path = 'W:\\csh\\Nursing\\Affiliation Agreements'
    #Read in term descriptions
    TermDescriptions = pd.read_excel('W:\\csh\\Nursing\\Schedules\\Term Descriptions.xlsx', header=0, converters={'Term':str})
    # Set path for attachment
    duties = 'W:\\csh\\Nursing\\Preceptor-Mentor Requests\\MENP\\Preceptor, Student, and Faculty Responsibilities for MENP Program.pdf'
    
    # Get coordinator's email
    recipient = row['Preceptor Email']
    # Write subject
    subject = 'DePaul University Preceptorship'
    # Write body   
    body = '''To: {0} {1}
    Re:	DePaul University Preceptorship

    Dear {2}:
    Thank you for agreeing to serve as a Preceptor for DePaul University student {3} {4} (the "Student"), for the time period {5}. 
    
    As a Preceptor you will provide a supervised learning experience for the Student, who will be simultaneously enrolled in the course NSG {6}-{7}: {8} (the "Course"). The Student's specific objectives for the learning experience and the Course will be jointly planned and approved by the University, the Preceptor, and the Student at the initiation of the learning experience. The learning experience must provide the Student with a minimum of {9} hours of direct, supervised clinical practice consistent with the State of Illinois Nurse Practice Act and in keeping with your typical responsibilities in your work setting. Evaluating the Student's achievement of the approved objectives will be a joint responsibility of the University, the Preceptor, and the Student at the conclusion of the learning experience.
    
    Your specific responsibilities as Preceptor are outlined in the attached document, entitled "Preceptor, Student, and Faculty Responsibilities."  We would also encourage you to review the Affiliation Agreement between DePaul University and your site for more information about the rights and responsibilities of various parties with respect to the placement. Please be sure to share this letter with your immediate supervisor. In the event that you cannot fulfill these responsibilities and your immediate supervisor assigns another preceptor for the Student, this letter applies to anyone who may be assigned as preceptor for the Student.
    
    I again thank you for agreeing to participate in this program and provide an invaluable learning experience for {10} {11}.  If you have any questions or concerns, please do not hesitate to contact me.
    
    Sincerely,
    
    Alexander Stachniak
    Coordinator of Data Management'''.format(row['Preceptor First Name'], row['Preceptor Last Name'], row['Preceptor First Name'], row['Student First Name'], row['Student Last Name'], row['Date'], row['Cr'], row['Sec'], row['Course Title'], row['Hours'], row['Student First Name'], row['Student Last Name'])
    
    # Get attachments    
    attachments = [duties]
    # Send email
    try:
        send_email(recipient, subject, body, attachments)
        # If successful, add email to list of jobs completed
        jobs_completed.append(recipient)
    except Exception as e:
        print('Exception occured for {0}: {1}'.format(recipient, e))
    
    # Map clinical site to official organization name
    # Gather sites and abbreviations
    site_dict = get_sites_from_ARC()
    # Convert to a list of all sites
    sites = list(site_dict.keys())
    # Gather dictionary of site abbreviations for storage/lookup
    abbr_dict = {}
    for site in sites:
        # Convert site name to an abbreviation
        abbr = re.sub(r'[^A-Z]', '', site)
        # If duplicate abbreviations exist, store as a list
        if abbr in abbr_dict:
            if type(abbr_dict[abbr]) == str:
                abbr_dict[abbr] = [abbr_dict[abbr]]
            abbr_dict[abbr].append(site)
        # Otherwise, store as a single string
        else:
            abbr_dict[abbr] = site
    abbr = list(abbr_dict.keys())
    # Test if org is in the site_dict
    if row['Clinical Site'] in site_dict.keys():
        org = site_dict[row['Clinical Site']]
    else:
        # Test if org is an abbreviation
        if re.match(r'^[A-Z]+$', row['Clinical Site']):
            org = find_best_string_match(row['Clinical Site'], abbr)
            # It is possible for multiple sites to share an abbreviation
            # If this happens, we can attempt to use the recipient email
            email_org = re.findall(r'@(\w+)\.', recipient)[0]
            org = find_best_string_match(email_org, org)
        else:
            org = find_best_string_match(row['Clinical Site'], sites)
    
    # Save a copy of the email
    save_path = ending_path + '\\' + org + '\\' + 'Preceptor Letters of Agreement' + '\\' + TermDescriptions.loc[TermDescriptions['Term'] == row['Term'], 'Academic Year'].item() + '\\' + TermDescriptions.loc[TermDescriptions['Term'] == row['Term'], 'Quarter'].item()   
    title = row['Preceptor Last Name'] + ', ' + row['Preceptor First Name']
    # Sleep for 1 second to allow time for Outlook to refresh
    sleep(1)
    save_email('Preceptor', save_path, title)

def test_path (save_path):
    '''Creates path. If exception = Path Exists, ignore. Otherwise raise.'''
    try:
        os.makedirs(save_path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

def save_email (folder, save_path, title):
    '''Saves the last email in a particular folder.
    Additional documentation: https://msdn.microsoft.com/en-us/library/microsoft.office.interop.outlook.mailitem_methods.aspx
    '''
    outlook = win32.Dispatch('outlook.application').GetNamespace("MAPI")
    inbox = outlook.Folders("astachn1@depaul.edu").Folders(folder)
    messages = inbox.Items
    message = messages.GetLast()
    test_path(save_path)
    message.SaveAs(save_path + '\\' + title + '.msg')
    
def get_sites_from_ARC ():
    '''A function that connects to the ARC Access database and grabs all
    site / corporation pairings.'''
    # Access Connection Strings
    access_conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=W:\\csh\\Nursing\\Agreement Request Center (ARC)\\Backend\\AffiliationAgreements_Backend.accdb;'
        )
    # Connect and open cursor
    cnxn = pyodbc.connect(access_conn_str)
    crsr = cnxn.cursor()
    # multiple join SQL statement
    sql = "SELECT c.[Site Name], d.Corporation FROM ((([Sites in Agreement] AS a) INNER JOIN Agreements AS b ON b.Agreement_ID = a.Agreement) INNER JOIN Sites AS c ON c.Site_ID = a.Site) INNER JOIN Corporations AS d ON d.Corporation_ID = b.Corporation"
    # Execute and save query results to dictionary
    crsr.execute(sql)
    site_dict = dict(crsr.fetchall())
    # Close connection
    crsr.close()
    cnxn.close()
    return site_dict

def save_excel_changes (file):
    '''Function saves changes made to excel file'''
    # Open excel workbook
    wb = openpyxl.load_workbook(file)
    # Open first sheet
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    # Gather column indices
    col_indices = {cell.value:n for n, cell in enumerate(list(sheet.rows)[0])}
    # Find column for Preceptor Email and Letter Sent
    colNum_email = col_indices['Preceptor Email'] + 1
    colNum_sent = col_indices['Preceptor Letter Sent'] + 1
    # Get today's date
    today = datetime.today().strftime('%m/%d/%Y')
    # Iterate through all rows in sheet, skipping first row (idx starts at 1)
    for rowNum in range(2, sheet.max_row):
        # Test if the email is in the jobs_completed list
        if sheet.cell(row=rowNum, column=colNum_email).value in jobs_completed:
            # Write today's date
            sheet.cell(row=rowNum, column=colNum_sent).value = today
            date_style = NamedStyle(name='datetime', number_format='MM/DD/YYY')
            sheet.cell(row=rowNum, column=colNum_sent).style = date_style
    # Save workbook
    wb.save(file)

#############################################################
# Main
#############################################################
@click.command()
@click.option(
        '--path',
        help='Folder path to current quarter, e.g. "2017-2018\Winter"',
)
def main(path):
    '''Main function.'''
    # Make a global list to store jobs completed
    global jobs_completed
    jobs_completed = []
    # Initial path
    starting_path = 'W:\\csh\\Nursing Administration\\Clinical Placement Files'
    quarters = ['Fall', 'Winter', 'Spring', 'Summer']
    # If user supplied a path, join it to the starting path
    if path:
        starting_path = os.path.join(starting_path, path)
    # Else, determine the correct path
    else:
        # Get all directories that match regex
        subfolders = [f.name for f in os.scandir(starting_path) if f.is_dir() and re.match(r'\d{4}-\d{4}', f.name)]
        # Sort the directories
        subfolders.sort()
        # Assume correct academic year is the last one
        ay = subfolders[-1]
        starting_path = os.path.join(starting_path, ay)
        # Now search for correct quarter
        subfolders = [f.name for f in os.scandir(starting_path) if f.is_dir()]
        # Assume last quarter is correct one
        for q in reversed(quarters):
            if q in subfolders:
                break
        starting_path = os.path.join(starting_path, q)
    
    # Read in latest preceptors file
    file = get_latest(starting_path)
    preceptors = pd.read_excel(file, header=0, converters={'Term':str, 'Cr':str, 'Empl ID':str, 'Hours':str})
    # Drop any rows without a preceptor email
    preceptors.dropna(subset=['Preceptor Email'], inplace=True)
    # Drop any rows with "TBD" as preceptor email
    preceptors = preceptors[~preceptors['Preceptor Email'].isin(['TBD', 'TBA'])]
    # Drop any rows where preceptor email was already sent or will be sent by RFU
    preceptors = preceptors[preceptors['Preceptor Letter Sent'].isnull()]
    # Email the preceptors
    preceptors.apply(email_preceptor, axis=1)
    # Save the current date in Excel file for preceptor's emailed
    save_excel_changes (file)

if __name__ == "__main__":
    main()
